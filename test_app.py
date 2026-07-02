import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core import DataError, export_summary, summarize_excel


SHEET_NAME = "上海人寿五个卡类型人员统计"


def make_workbook(path, rows, sheet_name=SHEET_NAME, headers=("等级", "生效", "过犹")):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


class StatisticsTest(unittest.TestCase):
    def test_groups_by_month_and_level(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            make_workbook(
                path,
                [
                    ("A", "", "保单过犹日期：2025-06-19"),
                    ("B", "", "保单过犹日期：2025-06-01"),
                    ("A", "", "保单过犹日期：2025-05-31"),
                ],
            )
            result = summarize_excel(path)
            self.assertEqual([row["月份"] for row in result], ["2025-05", "2025-06"])
            self.assertEqual(result[0]["A"], 1)
            self.assertEqual(result[1]["A"], 1)
            self.assertEqual(result[1]["B"], 1)
            self.assertEqual(sum(row["合计"] for row in result), 3)

    def test_rejects_missing_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            make_workbook(path, [], sheet_name="其他工作表")
            with self.assertRaisesRegex(DataError, "缺少工作表"):
                summarize_excel(path)

    def test_rejects_wrong_header(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            make_workbook(path, [], headers=("卡等级", "生效", "过犹"))
            with self.assertRaisesRegex(DataError, "模板表头错误"):
                summarize_excel(path)

    def test_rejects_invalid_level(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            make_workbook(path, [("F", "", "保单过犹日期：2025-06-19")])
            with self.assertRaisesRegex(DataError, "第 2 行等级无效"):
                summarize_excel(path)

    def test_rejects_invalid_date(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            make_workbook(path, [("A", "", "2025/06/19")])
            with self.assertRaisesRegex(DataError, "第 2 行过犹日期无效"):
                summarize_excel(path)

    def test_exports_integer_summary(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "output.xlsx"
            rows = [{"月份": "2025-06", "A": 2, "B": 1, "C": 0, "D": 0, "E": 0, "合计": 3}]
            export_summary(rows, path)
            workbook = load_workbook(path, data_only=True)
            sheet = workbook["月度等级统计"]
            self.assertEqual([cell.value for cell in sheet[1]], ["月份", "A", "B", "C", "D", "E", "合计"])
            self.assertEqual([cell.value for cell in sheet[2]], ["2025-06", 2, 1, 0, 0, 0, 3])
            workbook.close()


if __name__ == "__main__":
    unittest.main()

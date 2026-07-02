import re
from collections import defaultdict
from datetime import date
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


SHEET_NAME = "上海人寿五个卡类型人员统计"
EXPECTED_HEADERS = {"A1": "等级", "C1": "过犹"}
LEVELS = ("A", "B", "C", "D", "E")
DATE_PATTERN = re.compile(r"保单过犹(?:⽇|日)期[：:]\s*(\d{4})-(\d{2})-(\d{2})")


class DataError(ValueError):
    """Excel 内容不符合固定模板。"""


def summarize_excel(path):
    suffix = Path(path).suffix.lower()
    if suffix not in {".xls", ".xlsx"}:
        raise DataError("仅支持 .xls 或 .xlsx 格式的 Excel 文件。")

    if suffix == ".xls":
        return _summarize_xls(path)
    return _summarize_xlsx(path)


def _build_summary(sheet_name, header_values, data_rows):
    if sheet_name != SHEET_NAME:
        raise DataError(f"缺少工作表“{SHEET_NAME}”。")

    for cell, expected in EXPECTED_HEADERS.items():
        actual = header_values[cell]
        if actual != expected:
            raise DataError(
                f"模板表头错误：{cell} 应为“{expected}”，实际为“{actual}”。"
            )

    counts = defaultdict(lambda: {level: 0 for level in LEVELS})
    row_count = 0
    for row_number, level_value, date_value in data_rows:
        if level_value is None and date_value is None:
            continue

        level = str(level_value).strip().upper() if level_value is not None else ""
        if level not in LEVELS:
            raise DataError(
                f"第 {row_number} 行等级无效：应为 A、B、C、D、E，实际为“{level_value}”。"
            )

        match = DATE_PATTERN.fullmatch(str(date_value).strip()) if date_value is not None else None
        if not match:
            raise DataError(
                f"第 {row_number} 行过犹日期无效：应为“保单过犹日期：YYYY-MM-DD”，"
                f"实际为“{date_value}”。"
            )

        year, month, day = map(int, match.groups())
        try:
            date(year, month, day)
        except ValueError as exc:
            raise DataError(f"第 {row_number} 行过犹日期不是有效日期：{date_value}。") from exc

        counts[f"{year:04d}-{month:02d}"][level] += 1
        row_count += 1

    if row_count == 0:
        raise DataError("工作表中没有可统计的数据。")

    return [
        {"月份": month, **counts[month], "合计": sum(counts[month].values())}
        for month in sorted(counts)
    ]


def _summarize_xlsx(path):
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise DataError(f"无法打开 Excel 文件：{exc}") from exc

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise DataError(f"缺少工作表“{SHEET_NAME}”。")
        sheet = workbook[SHEET_NAME]
        rows = (
            (row_number, level_value, date_value)
            for row_number, (level_value, _effective_value, date_value) in enumerate(
                sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True), start=2
            )
        )
        return _build_summary(
            sheet.title,
            {cell: sheet[cell].value for cell in EXPECTED_HEADERS},
            rows,
        )
    finally:
        workbook.close()


def _summarize_xls(path):
    try:
        workbook = xlrd.open_workbook(path, on_demand=True)
    except Exception as exc:
        raise DataError(f"无法打开 Excel 文件：{exc}") from exc

    try:
        if SHEET_NAME not in workbook.sheet_names():
            raise DataError(f"缺少工作表“{SHEET_NAME}”。")
        sheet = workbook.sheet_by_name(SHEET_NAME)

        def value(row_index, column_index):
            if row_index >= sheet.nrows or column_index >= sheet.ncols:
                return None
            cell_value = sheet.cell_value(row_index, column_index)
            return cell_value if cell_value != "" else None

        rows = (
            (row_index + 1, value(row_index, 0), value(row_index, 2))
            for row_index in range(1, sheet.nrows)
        )
        return _build_summary(
            sheet.name,
            {"A1": value(0, 0), "C1": value(0, 2)},
            rows,
        )
    finally:
        workbook.release_resources()


def export_summary(rows, path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "月度等级统计"
    headers = ("月份", *LEVELS, "合计")
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 14
    for column in "BCDEFG":
        sheet.column_dimensions[column].width = 11
        for cell in sheet[column]:
            cell.alignment = Alignment(horizontal="center")

    workbook.save(path)
    workbook.close()
